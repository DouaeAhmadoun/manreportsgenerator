# =============================================================================
# excel_utils.py - Utilitaires Excel
# =============================================================================

import pandas as pd
import openpyxl
import re
import streamlit as st
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from .validation import ExcelContentValidators, ExcelMappingValidators, ExcelDataValidators


@dataclass
class ExcelAnalysisResult:
    """Résultat de l'analyse d'un fichier Excel"""
    success: bool
    sheets: List[str]
    recommended_sheet: Optional[str]
    confidence_score: int
    red_rows_detected: int
    error: Optional[str] = None


@dataclass
class SimulationData:
    """Données d'une simulation extraite d'Excel"""
    id: int
    numero_essai_original: str
    navire: str
    etat_chargement: str
    manoeuvre: str
    conditions_env: Dict[str, str]
    resultat: str
    commentaire_pilote: str
    remorqueurs: str = ""
    poste: str = ""
    pilote: str = ""
    bord: str = ""
    entree: str = ""
    condition: str = ""
    images: List[Dict] = None
    is_emergency_scenario: bool = False
    row_color: str = "normal"
    
    def __post_init__(self):
        if self.images is None:
            self.images = []


class ExcelColumnMapper:
    """Mappeur de colonnes Excel vers les champs de l'application"""
    
    COLUMN_MAPPING = {
        "numero_essai": [
            "N° Essai", "N° d'essai", "N°", "Test N°", "Essai N°", 
            "Numero", "Number", "No", "Num", "ID", "Test", "Essai",
            "N°Essai", "NumEssai", "Trial", "Trial N°", "Trial No"
        ],
        "condition": [  # ✨ NOUVEAU CHAMP AJOUTÉ
            "Condition", "Conditions", "État", "Etat", "Status", 
            "State", "Condition générale", "Condition spéciale",
            "Contexte", "Situation", "Circonstances"
        ],
        "vent": ["Vent", "vent", "wind", "Wind"],
        "houle": ["Houle", "Houle ", "Wave"],
        "courant": ["Courant", "Courant ", "Current"],
        "maree": ["Marée", "Marée", "Tide"],
        "manoeuvre": ["Manœuvre", "Manoeuvre", "Maneuver"],
        "navire": ["Navire"],
        "etat_chargement": ["Etat de chargement"],
        "remorqueurs": ["Remorqueurs", "Tugs"],
        "poste": ["Poste", "Poste ", "Berth"],
        "pilote": ["Pilote/Instructeur", "Pilote"],
        "commentaire": ["Commentaire", "Commentaires", "comments", "Comments"],
        "bord": ["Unnamed", "Bord", "Côté", "Side"],
        "entree": ["Entrance", "Entrée", "Entry", "Entrance "]
    }


def map_columns(self, available_columns: List[str]) -> Dict[str, str]:
        """Mappe les colonnes disponibles aux champs de l'application"""
        mapped_columns = {}
        
        for key, possible_names in self.COLUMN_MAPPING.items():
            for name in possible_names:
                if name in available_columns:
                    mapped_columns[key] = name
                    break
            
            # Détection intelligente pour le numéro d'essai
            if key == "numero_essai" and key not in mapped_columns:
                potential_cols = self.find_potential_numero_columns(available_columns)
                if potential_cols:
                    mapped_columns[key] = potential_cols[0]
                    st.info(f"💡 **Colonne numéro d'essai détectée automatiquement:** `{potential_cols[0]}`")
        
        return mapped_columns
    
def find_potential_numero_columns(self, columns: List[str]) -> List[str]:
    """Trouve les colonnes potentielles pour le numéro d'essai"""
    potential_cols = []
    keywords = ["n°", "num", "test", "essai", "trial", "id"]
        
    for col in columns:
        col_lower = str(col).lower()
        if any(keyword in col_lower for keyword in keywords):
            potential_cols.append(col)
        
    return potential_cols


class ExcelColorDetector:
    """Détecteur de couleurs dans les cellules Excel"""
    
    @staticmethod
    def is_color_red(color_value) -> bool:
        """Détermine si une couleur est rouge"""
        try:
            # Gérer différents formats de couleur
            if hasattr(color_value, 'rgb'):
                hex_color = color_value.rgb
            elif hasattr(color_value, 'value'):
                hex_color = color_value.value
            else:
                hex_color = str(color_value)
            
            # Convertir en string si nécessaire
            if not isinstance(hex_color, str):
                hex_color = str(hex_color)
            
            # Ignorer les couleurs par défaut/transparentes
            if not hex_color or hex_color in ['00000000', 'None', 'auto']:
                return False
            
            # Nettoyer la couleur hex
            clean_hex = hex_color.upper()
            
            # Enlever le préfixe alpha si présent (8 caractères -> 6)
            if len(clean_hex) == 8 and clean_hex.startswith('FF'):
                clean_hex = clean_hex[2:]
            
            # Vérifier qu'on a bien 6 caractères hex
            if len(clean_hex) != 6:
                return False
            
            # Convertir en RGB
            try:
                r = int(clean_hex[0:2], 16)
                g = int(clean_hex[2:4], 16) 
                b = int(clean_hex[4:6], 16)
                
                # Critères pour détecter le rouge
                return r > g and r > b and r > 150
                
            except ValueError:
                return False
                
        except Exception:
            return False
    
    def detect_red_rows(self, worksheet, max_row: int, max_col: int = 15) -> set:
        """Détecte les lignes avec des cellules rouges"""
        red_rows = set()
        
        for row_idx in range(2, max_row + 1):  # Commencer à ligne 2
            for col_idx in range(1, min(max_col + 1, worksheet.max_column + 1)):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                if self._is_cell_red(cell):
                    red_rows.add(row_idx - 2)  # -2 pour ajuster avec l'index pandas
                    break
        
        return red_rows
    
    def _is_cell_red(self, cell) -> bool:
        """Vérifie si une cellule est rouge"""
        # Méthode 1: Police rouge
        if cell.font and cell.font.color:
            if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                if self.is_color_red(cell.font.color.rgb):
                    return True
        
        # Méthode 2: Fond rouge
        if cell.fill and cell.fill.start_color:
            if hasattr(cell.fill.start_color, 'rgb') and cell.fill.start_color.rgb:
                if self.is_color_red(cell.fill.start_color.rgb):
                    return True
        
        # Méthode 3: fgColor
        if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
            if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                if self.is_color_red(cell.fill.fgColor.rgb):
                    return True
        
        return False


class ExcelTextCleaner:
    """Nettoyeur de texte Excel"""
    
    @staticmethod
    def clean_text(text) -> str:
        """Nettoie le texte importé depuis Excel"""
        if not isinstance(text, str):
            return str(text) if text is not None else ""
        
        # Nettoyer les caractères spéciaux
        cleaned = text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
        
        # Supprimer les espaces multiples
        cleaned = re.sub(r'\s+', ' ', cleaned)
        
        # Supprimer les espaces en début/fin
        return cleaned.strip()
    
    @staticmethod
    def normalize_trial_number(numero_essai) -> str:
        """Normalise un numéro d'essai"""
        if not numero_essai:
            return ""
        
        numero_str = str(numero_essai).strip()
        
        if not numero_str:
            return ""
        
        # Si c'est un nombre pur (12.0), garder l'entier
        try:
            float_val = float(numero_str)
            if float_val.is_integer():
                return str(int(float_val))
        except:
            pass
        
        # Extraire avant tiret/parenthèse
        match = re.match(r'^([^-()]+)', numero_str)
        if match:
            numero_str = match.group(1).strip()
        
        # Normaliser bis/ter: "12 bis" → "12bis"
        numero_str = re.sub(r'\s+(bis|ter|quater)\b', r'\1', numero_str, flags=re.IGNORECASE)
        
        # Supprimer non-alphanumériques + minuscules
        numero_clean = re.sub(r'[^a-zA-Z0-9]', '', numero_str).lower()
        
        return numero_clean


class ExcelSuccessDetector:
    """Détecteur de succès/échec dans les commentaires"""
    
    SUCCESS_KEYWORDS = [
        "concluant", "réussi", "faisable", "bon déroulement", 
        "sans difficulté", "maîtrisé", "succès", "réussite",
        "satisfaisant", "correct", "acceptable"
    ]
    
    FAILURE_KEYWORDS = [
        "échoué", "échec", "difficulté", "touché", "impossible", 
        "abandon", "problème", "erreur", "collision", "incident",
        "non réalisable", "dangereux"
    ]
    
    def detect_result(self, comment: str) -> str:
        """Détecte automatiquement le succès/échec d'une simulation"""
        if not comment:
            return "Non défini"
        
        comment_lower = comment.lower()
        
        # Détection basée sur des phrases spécifiques
        if "manœuvre échouée" in comment_lower or "manoeuvre échouée" in comment_lower:
            return "Échec"
        if "manœuvre concluante" in comment_lower or "manoeuvre concluante" in comment_lower:
            return "Réussite"
        if "bon déroulement" in comment_lower:
            return "Réussite"
        
        # Compter les occurrences
        success_count = sum(1 for word in self.SUCCESS_KEYWORDS if word in comment_lower)
        failure_count = sum(1 for word in self.FAILURE_KEYWORDS if word in comment_lower)
        
        # Décision basée sur le score
        if success_count > failure_count:
            return "Réussite"
        elif failure_count > success_count:
            return "Échec"
        else:
            return "Non défini"


class ExcelProcessor:
    def __init__(self):
        self.column_mapper = ExcelColumnMapper()
        self.color_detector = ExcelColorDetector()
        self.text_cleaner = ExcelTextCleaner()
        self.success_detector = ExcelSuccessDetector()
        
        self.smart_detector = SmartExcelColumnDetector()
        self.mapping_interface = ExcelMappingInterface()
    
    def analyze_excel_file(self, excel_file) -> ExcelAnalysisResult:
        """Analyse la structure d'un fichier Excel"""
        try:
            # Analyser avec pandas
            pd = get_pandas()
            excel_data = pd.read_excel(excel_file, sheet_name=None, dtype=str)
            
            # Analyser avec openpyxl pour les couleurs
            wb = openpyxl.load_workbook(excel_file, data_only=False)
            
            sheets = list(excel_data.keys())
            confidence_score = 0
            recommended_sheet = None
            red_rows_total = 0
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                df = excel_data[sheet_name]
                
                # Calculer score de confiance
                score = self._calculate_confidence_score(df.columns)
                
                if score > confidence_score:
                    confidence_score = score
                    recommended_sheet = sheet_name
                
                # Détecter les lignes rouges
                red_rows = self.color_detector.detect_red_rows(ws, ws.max_row)
                red_rows_total += len(red_rows)
            
            return ExcelAnalysisResult(
                success=True,
                sheets=sheets,
                recommended_sheet=recommended_sheet,
                confidence_score=confidence_score,
                red_rows_detected=red_rows_total
            )
            
        except Exception as e:
            return ExcelAnalysisResult(
                success=False,
                sheets=[],
                recommended_sheet=None,
                confidence_score=0,
                red_rows_detected=0,
                error=str(e)
            )
    
    def _calculate_confidence_score(self, columns) -> int:
        """Calcule un score de confiance pour une feuille Excel"""
        expected_keywords = [
            "essai", "navire", "manoeuvre", "vent", "houle", 
            "commentaire", "pilote", "remorqueur"
        ]
        
        score = 0
        for col in columns:
            col_lower = str(col).lower()
            for keyword in expected_keywords:
                if keyword in col_lower:
                    score += 1
                    break
        
        return score
    
    def _extract_field_value(self, row, mapped_columns: Dict[str, str], field: str) -> str:
        """Extrait et nettoie la valeur d'un champ"""
        if field not in mapped_columns:
            return ""
        
        excel_col = mapped_columns[field]
        value = row.get(excel_col, "")
        
        return self.text_cleaner.clean_text(value)
    
EnhancedExcelProcessor = ExcelProcessor


# =============================================================================
# Système Excel adaptatif intelligent
# =============================================================================

@dataclass
class ColumnMappingResult:
    """Résultat du mapping intelligent de colonnes"""
    mapped_columns: Dict[str, str]  # champ_app -> nom_colonne_excel
    confidence_scores: Dict[str, float]  # confiance pour chaque mapping
    unmapped_columns: List[str]  # colonnes Excel non mappées
    missing_required: List[str]  # champs requis manquants
    suggested_mappings: Dict[str, List[str]]  # suggestions pour colonnes ambiguës
    warnings: List[str]
    errors: List[str]


@dataclass
class ColumnPattern:
    """Pattern de reconnaissance pour une colonne"""
    field_name: str
    keywords: List[str]  # Mots-clés à rechercher
    content_patterns: List[str]  # Patterns regex pour le contenu
    validation_func: Optional[callable] = None  # Fonction de validation du contenu
    required: bool = False
    weight: float = 1.0  # Poids pour le scoring

class SmartExcelColumnDetector:    
    def __init__(self):
        self.patterns = self._create_column_patterns()
        self.content_validators = self._create_content_validators()
        self.debug_mode = True  # Active le debug
    
    def analyze_excel_structure(self, df) -> ColumnMappingResult:
        """Analyse avec debug détaillé"""
                
        # Initialiser le résultat
        result = ColumnMappingResult(
            mapped_columns={},
            confidence_scores={},
            unmapped_columns=[],
            missing_required=[],
            suggested_mappings={},
            warnings=[],
            errors=[]
        )
        
        excel_columns = df.columns.tolist()
        st.write(f"📊 Colonnes Excel trouvées: {excel_columns}")
        
        # Étape 1: Scoring des colonnes avec debug
        column_scores = self._score_all_columns(df, excel_columns)
        
        # Étape 2: Afficher le scoring pour debug
        #self._display_scoring_debug(column_scores)
        
        # Étape 3: Mapping optimal avec debug
        mapping = self._find_optimal_mapping(column_scores)
        
        # Étape 4: Validation avec debug
        validated_mapping = self._validate_mapping(df, mapping, result)
        
        # Étape 5: Identifier les problèmes
        self._identify_issues(excel_columns, validated_mapping, result)
        
        return result
    
    def _create_column_patterns(self) -> List[ColumnPattern]:
        """Crée les patterns de reconnaissance des colonnes"""
        return [
            # Numéro d'essai - TRÈS SPÉCIFIQUE ET PRIORITAIRE
            ColumnPattern(
                field_name="numero_essai",
                keywords=[
                    "test", "essai", "trial", "n°", "num", "numero", "number", 
                    "id", "index", "run", "simulation", "exp", "experience",
                    "#", "no", "nr", "série", "serie", "seq", "sequence"
                ],
                content_patterns=[
                    r'^\d+$',  # Nombres simples: 1, 2, 3
                    r'^\d+[a-z]*$',  # Nombres avec suffixe: 12bis, 15ter
                    r'^[a-z]\d+$',  # Lettres + nombres: E1, T12, S5
                    r'^\d+[-_]\d+$',  # Nombres composés: 12-1, 15_2
                    r'^\d+[.,]\d+$',  # Nombres décimaux: 1.1, 2,5
                ],
                required=True,
                weight=3.0  # POIDS MAXIMUM - CHAMP OBLIGATOIRE
            ),
            
            # Commentaires - TRÈS IMPORTANT POUR ÉVITER LES CONFUSIONS
            ColumnPattern(
                field_name="commentaire",
                keywords=[
                    "comment", "comments", "commentaire", "commentaires",
                    "observation", "observations", "note", "notes", 
                    "remark", "remarks", "description", "descriptions",
                    "detail", "details", "outcome", "result_comment"
                ],
                content_patterns=[
                    r'.{20,}',  # Texte long (commentaires sont généralement longs)
                    r'.*\b(manœuvre|manoeuvre|simulation|pilote|remorqueur|navire)\b.*',  # Vocabulaire du domaine
                    r'.*\b(réussi|échec|difficulté|problème|bon|mauvais)\b.*',  # Mots d'évaluation
                    r'.*\b(accost|depart|approach|berth)\b.*',  # Mots de manœuvre
                ],
                weight=2.5,
                validation_func=lambda series: len([s for s in series.dropna().astype(str).head(10) if len(s) > 15]) > len(series.dropna().head(10)) * 0.6
            ),
            
            # Conditions météorologiques - PATTERNS SOPHISTIQUÉS
            ColumnPattern(
                field_name="vent",
                keywords=[
                    "vent", "wind", "viento", "vento", "breeze", "gale",
                    "air", "atmosphere", "weather_wind"
                ],
                content_patterns=[
                    r'\d+\s*(kts?|knots?|nds?|nœuds?|kt|mph|km/h)',  # Vitesse: 25 kts, 30 mph
                    r'[NSEW]{1,3}\s*\d+',  # Direction + vitesse: NW 25, ENE 15
                    r'\d+\s*[NSEW]{1,3}',  # Vitesse + direction: 25 NW
                    r'calm|calme|nul|null|negligible|light',  # Conditions calmes
                    r'\d+°',  # Direction en degrés: 245°
                ],
                weight=2.0,
                validation_func=lambda series: len([s for s in series.dropna().astype(str).head(10) 
                                                if any(w in s.lower() for w in ['kt', 'wind', 'calm', 'nw', 'ne', 'sw', 'se'])]) > 0
            ),
            
            ColumnPattern(
                field_name="houle", 
                keywords=[
                    "houle", "wave", "waves", "swell", "vague", "vagues",
                    "sea", "mer", "ondulation", "surf"
                ],
                content_patterns=[
                    r'\d+[.,]?\d*\s*m',  # Hauteur: 2.5m, 3,2m, 4m
                    r'\d+\s*[.,]?\d*[/\s]*\d+\s*s',  # Hauteur/période: 2.5/12s, 3,2 10s
                    r'[NSEW]{1,3}[/\s]*\d+',  # Direction/hauteur: NW/2.5m, N 3m
                    r'\d+\s*(sec|second|periode|period)',  # Période: 12 sec
                ],
                weight=2.0,
                validation_func=lambda series: len([s for s in series.dropna().astype(str).head(10) 
                                                if any(w in s.lower() for w in ['m', 's', 'wave', 'houle'])]) > 0
            ),
            
            ColumnPattern(
                field_name="courant",
                keywords=[
                    "courant", "current", "corriente", "deriva", "drift",
                    "flow", "stream", "circulation"
                ],
                content_patterns=[
                    r'\d+[.,]?\d*\s*(kts?|kt|nds?|knots?)',  # Vitesse courant: 1.2 kts
                    r'[NSEW]{1,3}\s*\d+',  # Direction: SE 1.5, N 0.8
                    r'nul|null|negligible|faible|nil|zero',  # Courant nul
                    r'\d+°',  # Direction en degrés
                    r'flood|ebb|flot|jusant',  # Types de courant de marée
                ],
                weight=2.0,
                validation_func=lambda series: len([s for s in series.dropna().astype(str).head(10) 
                                                if any(w in s.lower() for w in ['kt', 'current', 'nul', 'flow'])]) > 0
            ),
            
            ColumnPattern(
                field_name="maree",
                keywords=[
                    "marée", "maree", "tide", "marea", "level", "hauteur",
                    "water_level", "niveau", "height", "elevation"
                ],
                content_patterns=[
                    r'[+\-]?\d+[.,]?\d*\s*m',  # Hauteur: +2.14m, -0.5m
                    r'[+\-]?\d+[.,]?\d*',  # Hauteur sans unité: +2.14, -0.5
                    r'PM|BM|HW|LW|HT|LT',  # Pleine mer, Basse mer, High Water, Low Water
                    r'ZH|CM|CD|MSL',  # Références altimétriques
                    r'flood|ebb|rising|falling',  # États de marée
                ],
                weight=1.8
            ),
            
            # Informations navire et manœuvre
            ColumnPattern(
                field_name="navire",
                keywords=[
                    "navire", "ship", "vessel", "bateau", "buque", "embarcación",
                    "containership", "container", "tanker", "bulk", "cargo"
                ],
                content_patterns=[
                    r'.{3,}',  # Au moins 3 caractères (noms de navires)
                    r'.*ship.*|.*vessel.*|.*carrier.*',  # Types de navires
                ],
                weight=1.8,
                validation_func=lambda series: len([s for s in series.dropna().astype(str).head(10) 
                                                if len(s) > 2 and not s.isdigit()]) > len(series.dropna().head(10)) * 0.7
            ),
            
            ColumnPattern(
                field_name="manoeuvre",
                keywords=[
                    "manœuvre", "manoeuvre", "maneuver", "maniobra", 
                    "operation", "mouvement", "approach", "departure",
                    "berthing", "unberthing", "docking", "undocking"
                ],
                content_patterns=[
                    r'.*accost.*|.*berth.*|.*atraque.*|.*dock.*',  # Accostage
                    r'.*depart.*|.*sortie.*|.*salida.*|.*undock.*',  # Appareillage
                    r'.*evit.*|.*avoid.*|.*emergency.*|.*urgence.*',  # Évitement
                    r'.*approach.*|.*approche.*',  # Approche
                ],
                weight=1.8
            ),
            
            # État de charge
            ColumnPattern(
                field_name="etat_chargement",
                keywords=[
                    "charge", "chargement", "load", "laden", "ballast", 
                    "light", "empty", "état", "etat", "condition",
                    "loading", "cargo", "displacement"
                ],
                content_patterns=[
                    r'.*charg.*|.*load.*|.*laden.*|.*full.*',  # Chargé
                    r'.*lège.*|.*light.*|.*empty.*|.*vide.*',  # À vide
                    r'.*ballast.*|.*lest.*',  # Ballast
                    r'.*partial.*|.*partiel.*',  # Partiellement chargé
                ],
                weight=1.5
            ),
            
            # Assistance remorquage
            ColumnPattern(
                field_name="remorqueurs",
                keywords=[
                    "remorqueur", "remorqueurs", "remolcador", "tug", "tugboat", "tugs",
                    "assistance", "aide", "help", "support", "towboat"
                ],
                content_patterns=[
                    r'\d+\s*(remorqueur|tug|remolcador)',  # Nombre de remorqueurs
                    r'.*ASD.*|.*tractor.*|.*conventional.*',  # Types de remorqueurs
                    r'.*assistance.*|.*aide.*|.*help.*',  # Assistance
                ],
                weight=1.5
            ),
            
            # Infrastructure portuaire
            ColumnPattern(
                field_name="poste",
                keywords=[
                    "poste", "berth", "quai", "wharf", "pier", "terminal",
                    "dock", "jetty", "position", "mooring"
                ],
                content_patterns=[
                    r'[A-Z]?\d+',  # P1, Q2, 15, B3
                    r'[A-Z]{1,3}\d*',  # TM1, PPR, CT
                    r'.*terminal.*|.*quai.*|.*berth.*',  # Noms de postes
                ],
                weight=1.2
            ),
            
            ColumnPattern(
                field_name="bord",
                keywords=[
                    "bord", "côté", "side", "lado", "port", "starboard", 
                    "babord", "tribord", "left", "right", "gauche", "droite"
                ],
                content_patterns=[
                    r'.*port.*|.*bâbord.*|.*babord.*|.*left.*|.*gauche.*',  # Bâbord
                    r'.*starboard.*|.*tribord.*|.*right.*|.*droite.*',  # Tribord
                    r'.*p.*|.*s.*',  # Abréviations P/S
                ],
                weight=1.0
            ),
            
            ColumnPattern(
                field_name="entree",
                keywords=[
                    "entrance", "entry", "entrée", "entree", "access", "accès",
                    "approach", "approche", "channel", "chenal", "gate"
                ],
                content_patterns=[
                    r'[A-Z]{2,4}\d*',  # TM1, TM2, MAIN, etc.
                    r'.*channel.*|.*chenal.*|.*entrance.*',  # Types d'entrées
                    r'.*main.*|.*principal.*|.*north.*|.*south.*',  # Directions
                ],
                weight=1.0
            ),
            
            # Personnel
            ColumnPattern(
                field_name="pilote",
                keywords=[
                    "pilote", "pilot", "instructeur", "instructor", "captain",
                    "capitaine", "master", "officer", "trainer"
                ],
                content_patterns=[
                    r'.+/.+',  # Format "Pilote/Instructeur"
                    r'.*pilot.*|.*instructeur.*|.*captain.*',  # Titres
                    r'[A-Z][a-z]+\s+[A-Z][a-z]+',  # Noms propres: Jean Dupont
                ],
                weight=1.0
            ),
            
            # Résultat et évaluation
            ColumnPattern(
                field_name="resultat",
                keywords=[
                    "résultat", "result", "outcome", "success", "échec", "reussite",
                    "evaluation", "assessment", "grade", "note", "score"
                ],
                content_patterns=[
                    r'^[01]$',  # 0 ou 1 binaire
                    r'.*success.*|.*réussi.*|.*concluant.*|.*pass.*',  # Succès
                    r'.*fail.*|.*échec.*|.*échoué.*|.*unsuccess.*',  # Échec
                    r'^(pass|fail|ok|nok)$',  # Résultats courts
                ],
                weight=1.8
            ),
            
            # Données temporelles
            ColumnPattern(
                field_name="date_simulation",
                keywords=[
                    "date", "day", "jour", "when", "quand", "time", "temps"
                ],
                content_patterns=[
                    r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',  # Dates: 25/12/2024
                    r'\d{4}-\d{2}-\d{2}',  # Format ISO: 2024-12-25
                    r'\d{1,2}/\d{1,2}/\d{4}',  # Format US: 12/25/2024
                ],
                weight=0.8
            ),
            
            ColumnPattern(
                field_name="heure_simulation",
                keywords=[
                    "heure", "hour", "time", "horaire", "timing", "when"
                ],
                content_patterns=[
                    r'\d{1,2}:\d{2}',  # Heures: 14:30, 9:15
                    r'\d{1,2}[h:]\d{2}',  # Formats français: 14h30
                    r'\d+\s*(min|minutes?|h|heures?)',  # Durées
                ],
                weight=0.8
            ),
            
            # Conditions spéciales
            ColumnPattern(
                field_name="visibilite",
                keywords=[
                    "visibility", "visibilité", "visibilite", "vue", "sight",
                    "fog", "brouillard", "clear", "clair"
                ],
                content_patterns=[
                    r'\d+\s*(m|km|nm|miles?)',  # Distances de visibilité
                    r'.*fog.*|.*brouillard.*|.*mist.*',  # Conditions dégradées
                    r'.*clear.*|.*good.*|.*excellent.*',  # Bonnes conditions
                ],
                weight=0.8
            ),
            
            # Données techniques
            ColumnPattern(
                field_name="vitesse",
                keywords=[
                    "speed", "vitesse", "velocity", "rate", "pace"
                ],
                content_patterns=[
                    r'\d+[.,]?\d*\s*(kt|kts|mph|km/h|nds?)',  # Vitesses
                    r'.*approach.*speed.*|.*vitesse.*approche.*',  # Vitesse d'approche
                ],
                weight=0.8
            ),
            
            ColumnPattern(
                field_name="distance",
                keywords=[
                    "distance", "range", "length", "longueur", "portée"
                ],
                content_patterns=[
                    r'\d+[.,]?\d*\s*(m|km|nm|ft|miles?)',  # Distances
                    r'.*stopping.*|.*arrêt.*|.*final.*',  # Distance d'arrêt
                ],
                weight=0.8
            )
        ]

    def _create_content_validators(self) -> Dict[str, callable]:
        """Crée les validateurs de contenu pour chaque type de champ"""
        return {
            "numero_essai": ExcelContentValidators.validate_numero_essai,
            "vent": ExcelContentValidators.validate_vent,
            "houle": ExcelContentValidators.validate_houle,
            "courant": ExcelContentValidators.validate_courant,
            "maree": ExcelContentValidators.validate_maree,
            "commentaire": ExcelContentValidators.validate_commentaire,
            "navire": ExcelContentValidators.validate_navire,
            "resultat": ExcelContentValidators.validate_resultat,
            "date_simulation": ExcelContentValidators.validate_date,
            "heure_simulation": ExcelContentValidators.validate_heure,
        }

    def _normalize_text(self, text: str) -> str:
        """Normalise un texte pour la comparaison"""
        if not text:
            return ""
        
        # Minuscules, supprimer accents et caractères spéciaux
        normalized = str(text).lower()
        
        # Remplacements spécifiques pour les accents
        replacements = {
            'é': 'e', 'è': 'e', 'ê': 'e', 'ë': 'e',
            'à': 'a', 'á': 'a', 'â': 'a', 'ä': 'a',
            'ù': 'u', 'ú': 'u', 'û': 'u', 'ü': 'u',
            'ò': 'o', 'ó': 'o', 'ô': 'o', 'ö': 'o',
            'ç': 'c', 'ñ': 'n',
            'œ': 'oe', 'æ': 'ae'
        }
        
        for accented, normal in replacements.items():
            normalized = normalized.replace(accented, normal)
        
        # Supprimer caractères non alphanumériques
        normalized = re.sub(r'[^a-z0-9]', '', normalized)
        
        return normalized

    def _fuzzy_match(self, str1: str, str2: str) -> float:
        """Calcul de similarité simple entre deux chaînes"""
        if not str1 or not str2:
            return 0.0
        
        # Calcul de distance de Levenshtein simplifiée
        len1, len2 = len(str1), len(str2)
        
        if len1 == 0:
            return 0.0
        if len2 == 0:
            return 0.0
        
        # Matrice de distance
        matrix = [[0] * (len2 + 1) for _ in range(len1 + 1)]
        
        for i in range(len1 + 1):
            matrix[i][0] = i
        for j in range(len2 + 1):
            matrix[0][j] = j
        
        for i in range(1, len1 + 1):
            for j in range(1, len2 + 1):
                cost = 0 if str1[i-1] == str2[j-1] else 1
                matrix[i][j] = min(
                    matrix[i-1][j] + 1,      # suppression
                    matrix[i][j-1] + 1,      # insertion
                    matrix[i-1][j-1] + cost  # substitution
                )
        
        distance = matrix[len1][len2]
        max_len = max(len1, len2)
        
        return 1.0 - (distance / max_len)

    def _calculate_column_score(self, df, excel_col: str, pattern: ColumnPattern) -> float:
        """Calcule le score de correspondance entre une colonne Excel et un pattern"""
        
        score = 0.0
        
        # 1. Score basé sur le nom de la colonne (50% du poids)
        name_score = self._score_column_name(excel_col, pattern.keywords)
        score += name_score * 0.5 * pattern.weight
        
        # 2. Score basé sur le contenu (50% du poids)
        content_score = self._score_column_content(df[excel_col], pattern.content_patterns)
        score += content_score * 0.5 * pattern.weight
        
        # 3. Bonus pour validation spécifique
        if pattern.field_name in self.content_validators:
            validator = self.content_validators[pattern.field_name]
            if self._test_content_validation(df[excel_col], validator):
                score += 0.2 * pattern.weight
        
        return min(score, 1.0)  # Cap à 1.0

    def _score_column_name(self, excel_col: str, keywords: List[str]) -> float:
        """Score la similarité entre le nom de colonne et les mots-clés"""
        
        col_normalized = self._normalize_text(excel_col)
        
        # Score exact match
        for keyword in keywords:
            keyword_normalized = self._normalize_text(keyword)
            if keyword_normalized == col_normalized:
                return 1.0
            if keyword_normalized in col_normalized:
                return 0.8
            if col_normalized in keyword_normalized:
                return 0.7
        
        # Score fuzzy match
        best_fuzzy = 0.0
        for keyword in keywords:
            fuzzy_score = self._fuzzy_match(col_normalized, self._normalize_text(keyword))
            best_fuzzy = max(best_fuzzy, fuzzy_score)
        
        return best_fuzzy

    def _score_column_content(self, series, patterns: List[str]) -> float:
        """Score le contenu d'une colonne selon des patterns"""
        
        if series.empty:
            return 0.0
        
        # Échantillonner les valeurs non nulles
        sample_values = series.dropna().astype(str).head(20).tolist()
        
        if not sample_values:
            return 0.0
        
        total_matches = 0
        
        for value in sample_values:
            value_normalized = str(value).strip().lower()
            
            for pattern in patterns:
                if re.search(pattern, value_normalized, re.IGNORECASE):
                    total_matches += 1
                    break  # Un seul match par valeur
        
        return total_matches / len(sample_values)

    def _find_optimal_mapping(self, column_scores: Dict[str, Dict[str, float]]) -> Dict[str, str]:
        """Trouve le mapping optimal (algorithme hongrois simplifié)"""
        
        mapping = {}
        used_columns = set()
        
        # Trier les champs par priorité (required d'abord, puis par score max)
        fields_priority = []
        for pattern in self.patterns:
            if pattern.field_name in column_scores:
                max_score = max(column_scores[pattern.field_name].values())
                priority = (pattern.required, max_score, pattern.weight)
                fields_priority.append((pattern.field_name, priority))
        
        # Trier par priorité décroissante
        fields_priority.sort(key=lambda x: x[1], reverse=True)
        
        # Assigner de manière gloutonne
        for field_name, _ in fields_priority:
            field_scores = column_scores[field_name]
            
            # Trouver la meilleure colonne non utilisée
            best_col = None
            best_score = 0
            
            for col, score in field_scores.items():
                if col not in used_columns and score > best_score:
                    # Validation supplémentaire avant assignation
                    if ExcelMappingValidators.validate_mapping_candidate(field_name, col, score):
                        best_col = col
                        best_score = score
            
            # Assigner si score suffisant
            if best_col and best_score >= 0.3:  # Seuil minimum
                mapping[field_name] = best_col
                used_columns.add(best_col)
        
        return mapping

    def _validate_mapping(self, df, mapping: Dict[str, str], result: ColumnMappingResult) -> Dict[str, str]:
        """Valide le mapping proposé"""
        
        validated_mapping = {}
        
        for field, excel_col in mapping.items():
            # Vérifier que la colonne existe toujours
            if excel_col not in df.columns:
                result.warnings.append(f"Colonne '{excel_col}' introuvable pour {field}")
                continue
            
            # Validation spécifique du contenu si disponible
            if field in self.content_validators:
                validator = self.content_validators[field]
                if not self._test_content_validation(df[excel_col], validator):
                    result.warnings.append(f"Contenu de '{excel_col}' suspect pour {field}")
            
            validated_mapping[field] = excel_col
            
            # Calculer le score de confiance final
            pattern = next((p for p in self.patterns if p.field_name == field), None)
            if pattern:
                final_score = self._calculate_column_score(df, excel_col, pattern)
                result.confidence_scores[field] = final_score
        
        result.mapped_columns = validated_mapping
        return validated_mapping

    def _identify_issues(self, excel_columns: List[str], mapping: Dict[str, str], result: ColumnMappingResult):
        """Identifie les problèmes et suggestions"""
        
        # Colonnes non mappées
        used_columns = set(mapping.values())
        result.unmapped_columns = [col for col in excel_columns if col not in used_columns]
        
        # Champs requis manquants
        required_fields = [p.field_name for p in self.patterns if p.required]
        result.missing_required = [field for field in required_fields if field not in mapping]
        
        # Suggestions pour colonnes ambiguës (faible confiance)
        for field, score in result.confidence_scores.items():
            if score < 0.6:  # Seuil de faible confiance
                # Chercher d'autres candidats possibles dans les colonnes non mappées
                candidates = []
                pattern = next((p for p in self.patterns if p.field_name == field), None)
                
                if pattern:
                    for col in result.unmapped_columns[:3]:  # Limiter à 3 suggestions
                        candidates.append(col)
                
                if candidates:
                    result.suggested_mappings[field] = candidates
        
        # Erreurs critiques
        if len(result.missing_required) > 0:
            result.errors.append(f"Champs requis manquants: {', '.join(result.missing_required)}")
        
        if len(result.mapped_columns) == 0:
            result.errors.append("Aucune colonne reconnue automatiquement")

    def _test_content_validation(self, series, validator: callable) -> bool:
        """Teste la validation du contenu"""
        from .validation import test_content_validation
        return test_content_validation(series, validator)

    def _score_all_columns(self, df, excel_columns: List[str]) -> Dict[str, Dict[str, float]]:
        """Score toutes les combinaisons colonne Excel <-> champ application"""
        
        scores = {}  # {field_name: {excel_col: score}}
        
        for pattern in self.patterns:
            field_scores = {}
            
            for excel_col in excel_columns:
                score = self._calculate_column_score(df, excel_col, pattern)
                if score > 0:
                    field_scores[excel_col] = score
            
            if field_scores:
                scores[pattern.field_name] = field_scores
        
        return scores

    def analyze_excel_structure(self, df) -> ColumnMappingResult:
        """Analyse la structure d'un DataFrame Excel et propose un mapping intelligent"""
        
        # Initialiser le résultat
        result = ColumnMappingResult(
            mapped_columns={},
            confidence_scores={},
            unmapped_columns=[],
            missing_required=[],
            suggested_mappings={},
            warnings=[],
            errors=[]
        )
        
        excel_columns = df.columns.tolist()
        
        # Étape 1: Scoring des colonnes
        column_scores = self._score_all_columns(df, excel_columns)
        
        # Étape 2: Mapping optimal
        mapping = self._find_optimal_mapping(column_scores)
        
        # Étape 3: Validation du mapping
        validated_mapping = self._validate_mapping(df, mapping, result)
        
        # Étape 4: Identifier les problèmes
        self._identify_issues(excel_columns, validated_mapping, result)
        
        return result



class ExcelMappingInterface:
    """Interface utilisateur pour le mapping intelligent des colonnes"""
    
    def __init__(self):
        self.detector = SmartExcelColumnDetector()
    
    def render_mapping_interface(self, df) -> Optional[Dict[str, str]]:
        """Rend l'interface de mapping et retourne le mapping validé - MÉTHODE PRINCIPALE"""
        
        st.subheader("🧠 Mapping intelligent des colonnes", divider=True)
        
        # Analyse automatique
        with st.spinner("Analyse intelligente du fichier Excel..."):
            analysis_result = self.detector.analyze_excel_structure(df)
        
        # Affichage des résultats
        self._display_analysis_results(analysis_result)
        
        # Interface de correction si nécessaire
        final_mapping = self._render_mapping_editor(df, analysis_result)
        
        return final_mapping
    
    def _display_analysis_results(self, result: ColumnMappingResult):
        """Affiche les résultats de l'analyse"""
        
        # Statut général
        if result.errors:
            st.error("❌ **Problèmes critiques détectés**")
            for error in result.errors:
                st.error(f"• {error}")
        elif result.warnings:
            st.warning("⚠️ **Mapping partiellement réussi avec avertissements**")
        else:
            st.success("✅ **Mapping automatique réussi**")
        
        # Colonnes mappées avec confiance
        if result.mapped_columns:
            st.write("**📋 Colonnes reconnues automatiquement :**")
            
            for field, excel_col in result.mapped_columns.items():
                confidence = result.confidence_scores.get(field, 0)
                confidence_pct = confidence * 100
                
                if confidence >= 0.8:
                    conf_color = "🟢"
                elif confidence >= 0.6:
                    conf_color = "🟡"
                else:
                    conf_color = "🔴"
                
                st.write(f"{conf_color} **{field}** ← `{excel_col}` (confiance: {confidence_pct:.0f}%)")
        
        # Avertissements
        if result.warnings:
            with st.expander("⚠️ Avertissements", expanded=True):
                for warning in result.warnings:
                    st.warning(f"• {warning}")
        
        # Colonnes non mappées
        if result.unmapped_columns:
            with st.expander(f"❓ Colonnes non reconnues ({len(result.unmapped_columns)})", expanded=False):
                for col in result.unmapped_columns:
                    st.write(f"• `{col}`")
        
        # Suggestions
        if result.suggested_mappings:
            with st.expander("💡 Suggestions de mapping", expanded=True):
                for field, suggestions in result.suggested_mappings.items():
                    st.write(f"**{field}** pourrait correspondre à : {', '.join(f'`{s}`' for s in suggestions)}")
    
    def _render_mapping_editor(self, df, result: ColumnMappingResult) -> Dict[str, str]:
        """Rend l'éditeur de mapping pour les corrections manuelles"""
        
        st.subheader("🔧 Vérification et correction du mapping", divider=True)
        
        # Interface de correction
        excel_columns = ["[Aucune]"] + df.columns.tolist()
        final_mapping = {}
        
        # Créer des colonnes pour l'interface
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.write("**🎯 Champs de l'application**")
        with col2:
            st.write("**📊 Colonnes Excel correspondantes**")
        
        # Rendre chaque champ avec sélecteur
        for pattern in self.detector.patterns:
            field = pattern.field_name
            field_label = field.replace('_', ' ').title()
            
            # Valeur par défaut (mapping automatique ou aucune)
            default_excel_col = result.mapped_columns.get(field, "[Aucune]")
            default_index = excel_columns.index(default_excel_col) if default_excel_col in excel_columns else 0
            
            col1, col2 = st.columns([1, 1])
            
            with col1:
                # Indicateur de statut
                if field in result.mapped_columns:
                    confidence = result.confidence_scores.get(field, 0)
                    if confidence >= 0.8:
                        status = "🟢"
                    elif confidence >= 0.6:
                        status = "🟡"
                    else:
                        status = "🔴"
                elif field in result.missing_required:
                    status = "❌"
                else:
                    status = "⚪"
                
                required_marker = " *" if pattern.required else ""
                st.write(f"{status} **{field_label}**{required_marker}")
                
                if pattern.required:
                    st.caption("Champ obligatoire")
            
            with col2:
                # Sélecteur de colonne Excel
                selected_col = st.selectbox(
                    f"Colonne pour {field}",
                    excel_columns,
                    index=default_index,
                    key=f"mapping_{field}",
                    label_visibility="collapsed"
                )
                
                # Ajouter au mapping final si pas "[Aucune]"
                if selected_col != "[Aucune]":
                    final_mapping[field] = selected_col
                
                # Affichage d'aperçu du contenu
                if selected_col != "[Aucune]" and selected_col in df.columns:
                    sample_values = df[selected_col].dropna().head(3).tolist()
                    if sample_values:
                        sample_text = ", ".join(str(v)[:20] for v in sample_values)
                        st.caption(f"Aperçu: {sample_text}...")
        
        # ✨ NOUVELLE SECTION: Colonnes non mappées
        unmapped_columns = [col for col in df.columns.tolist() 
                          if col not in final_mapping.values()]
        
        if unmapped_columns:
            st.write("---")
            st.subheader("➕ Colonnes non utilisées", divider=True)
            st.write("**💡 Voulez-vous ajouter certaines colonnes comme nouveaux champs ?**")
            
            # Interface pour ajouter les colonnes non mappées
            additional_fields = self._render_unmapped_columns_selector(df, unmapped_columns)
            
            # Ajouter les champs supplémentaires au mapping final
            final_mapping.update(additional_fields)
        
        # Validation du mapping final
        st.write("---")
        
        # Vérifier les champs requis
        required_fields = [p.field_name for p in self.detector.patterns if p.required]
        missing_required = [f for f in required_fields if f not in final_mapping]
        
        if missing_required:
            st.error(f"❌ **Champs obligatoires manquants :** {', '.join(missing_required)}")
            st.error("**Veuillez corriger le mapping avant de continuer.**")
            return None
        
        # Vérifier les doublons
        used_columns = list(final_mapping.values())
        duplicates = [col for col in used_columns if used_columns.count(col) > 1]
        
        if duplicates:
            st.error(f"❌ **Colonnes utilisées plusieurs fois :** {', '.join(set(duplicates))}")
            st.error("**Chaque colonne Excel ne peut être mappée qu'à un seul champ.**")
            return None
        
        # Aperçu du mapping final
        if final_mapping:
            st.success(f"✅ **Mapping validé : {len(final_mapping)} champs mappés**")
            
            # Séparer champs standards et nouveaux
            standard_fields = [p.field_name for p in self.detector.patterns]
            new_fields = [field for field in final_mapping.keys() if field not in standard_fields]
            
            with st.expander("📋 Résumé du mapping final", expanded=False):
                if any(field in standard_fields for field in final_mapping.keys()):
                    st.write("**🎯 Champs standards :**")
                    for field, excel_col in final_mapping.items():
                        if field in standard_fields:
                            st.write(f"• **{field}** ← `{excel_col}`")
                
                if new_fields:
                    st.write("**➕ Nouveaux champs ajoutés :**")
                    for field in new_fields:
                        excel_col = final_mapping[field]
                        st.write(f"• **{field}** ← `{excel_col}`")
        
        return final_mapping
    
    def _render_unmapped_columns_selector(self, df, unmapped_columns: List[str]) -> Dict[str, str]:
        """Interface pour sélectionner les colonnes non mappées à ajouter"""
        from .validation import normalize_field_name

        additional_fields = {}
        
        if not unmapped_columns:
            return additional_fields
        
        # Afficher les colonnes non mappées avec aperçu
        st.write(f"**📋 {len(unmapped_columns)} colonne(s) non utilisée(s) :**")
        
        for i, col in enumerate(unmapped_columns):
            with st.container():
                col1, col2, col3 = st.columns([1, 2, 1])
                
                with col1:
                    # Checkbox pour inclure cette colonne
                    include_col = st.checkbox(
                        f"Ajouter",
                        key=f"include_unmapped_{i}",
                        help=f"Ajouter la colonne '{col}' comme nouveau champ"
                    )
                
                with col2:
                    st.write(f"**`{col}`**")
                    
                    # Aperçu du contenu
                    if col in df.columns:
                        sample_values = df[col].dropna().head(3).tolist()
                        if sample_values:
                            sample_text = ", ".join(str(v)[:25] for v in sample_values)
                            st.caption(f"Contenu: {sample_text}...")
                        else:
                            st.caption("Colonne vide")
                
                with col3:
                    if include_col:
                        # Nom du champ personnalisé
                        field_name = st.text_input(
                            "Nom du champ",
                            value=normalize_field_name(col),
                            key=f"field_name_unmapped_{i}",
                            label_visibility="collapsed",
                            placeholder="nom_du_champ"
                        )
                        
                        # Validation du nom de champ
                        if field_name and ExcelMappingValidators.validate_field_name(field_name):
                            additional_fields[field_name] = col
                        elif field_name:
                            st.error("❌ Nom invalide")
        
        # Résumé des ajouts
        if additional_fields:
            st.success(f"✅ **{len(additional_fields)} nouveau(x) champ(s) à ajouter**")
            
            with st.expander("👁️ Aperçu des nouveaux champs", expanded=False):
                for field_name, excel_col in additional_fields.items():
                    # Type de données détecté
                    data_type = self._detect_data_type(df[excel_col])
                    st.write(f"• **{field_name}** ← `{excel_col}` ({data_type})")
        
        return additional_fields
    
    def _detect_data_type(self, series) -> str:
        """Détecte le type de données d'une série"""
        
        sample = series.dropna().head(20)
        
        if sample.empty:
            return "vide"
        
        # Convertir en strings pour analyse
        sample_str = sample.astype(str)
        
        # Vérifier si c'est numérique
        numeric_count = 0
        for value in sample_str:
            try:
                float(value.replace(',', '.'))
                numeric_count += 1
            except ValueError:
                pass
        
        if numeric_count > len(sample) * 0.8:
            return "numérique"
        
        # Vérifier si c'est des dates
        date_patterns = [
            r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',
            r'\d{4}-\d{2}-\d{2}',
            r'\d{1,2}/\d{1,2}/\d{4}'
        ]
        
        date_count = 0
        for value in sample_str:
            if any(re.search(pattern, value) for pattern in date_patterns):
                date_count += 1
        
        if date_count > len(sample) * 0.6:
            return "date"
        
        # Vérifier si c'est du temps/heure
        time_patterns = [
            r'\d{1,2}:\d{2}',
            r'\d+\s*(min|minutes?|h|heures?|sec|secondes?)'
        ]
        
        time_count = 0
        for value in sample_str:
            if any(re.search(pattern, value, re.IGNORECASE) for pattern in time_patterns):
                time_count += 1
        
        if time_count > len(sample) * 0.6:
            return "temps"
        
        # Vérifier si c'est catégoriel (peu de valeurs uniques)
        unique_values = len(sample.unique())
        if unique_values <= 10 and unique_values < len(sample) * 0.5:
            return "catégoriel"
        
        # Par défaut: texte
        return "texte"


def get_pandas():
    """Importe pandas (fonction de compatibilité)"""
    import pandas as pd
    return pd

def normalize_trial_number(numero_essai) -> str:
    """ Fonction de compatibilité """
    return ExcelTextCleaner.normalize_trial_number(numero_essai)
